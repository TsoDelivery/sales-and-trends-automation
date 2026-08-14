#!/usr/bin/env python3
"""Validate (and optionally correct) Sales & Trends store tabs against the P&L.

Pipeline: TRIS Accounting -> email -> Min's Drive folder -> this script ->
compare to the live Sales & Trends store tabs -> report, and with --write,
correct the cells the P&L is authoritative for.

Default is a DRY RUN. Nothing is written without --write.

GRAIN SAFETY (the important part)
---------------------------------
The store tabs are calendar-month grain. The P&L header carries its own date
range and store_pl.extract() REFUSES anything that is not a whole calendar
month. A 28-day fiscal-period export understates these rows by 10-27%.

AUTHORITY (Angell, 2026-08-14)
------------------------------
The P&L is the reconciled source of truth for the mapped channel columns,
including Uber Eats and DoorDash where it runs 2-11% below the sheet's previous
figures (the P&L nets discounts and refunds). --write therefore corrects those.

Carryout and Delivery are NOT written: the P&L has one combined "Total Grafana
Sales" line covering two sheet columns and cannot apportion between them. The
variance is reported so a person can decide.

Usage:
  python3 scripts/validate-store-pl.py                        # newest P&L in Drive, dry run
  python3 scripts/validate-store-pl.py --file pl.csv          # local file
  python3 scripts/validate-store-pl.py --drive-id <fileId>
  python3 scripts/validate-store-pl.py --write                 # apply corrections
  python3 scripts/validate-store-pl.py --json report.json
"""

import argparse
import csv
import datetime as dt
import json
import os
import re
import subprocess
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import sheets_io
import store_pl as sp

REPO = Path(__file__).resolve().parent.parent

# 2026 floor: earlier years are reconciled history and must never be rewritten.
MIN_YEAR = 2026


def run(cmd, timeout=300):
    p = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
    if p.returncode != 0:
        raise RuntimeError(f"{' '.join(cmd[:3])} failed: {(p.stderr or p.stdout)[:400]}")
    return p.stdout


# ------------------------------------------------------------------------ drive

def find_drive_pl(year=None, month=None):
    """Newest Drive file that looks like a monthly P&L export."""
    out = run(["gog", "drive", "search", "P&L", "--max", "40"])
    best = None
    for line in out.splitlines():
        parts = re.split(r"\s{2,}", line.strip())
        if len(parts) < 2 or parts[0] in ("ID", "") or line.startswith("#"):
            continue
        fid, name = parts[0], parts[1]
        m = re.match(r"P&L\s+([A-Za-z]+)[_ ](\d{4})", name)
        if not m:
            continue
        try:
            mo = dt.datetime.strptime(m.group(1)[:3], "%b").month
        except ValueError:
            continue
        yr = int(m.group(2))
        if year and month and (yr, mo) != (year, month):
            continue
        if best is None or (yr, mo) > best[2]:
            best = (fid, name, (yr, mo))
    return best


def download_pl(file_id, dest_dir):
    dest = Path(dest_dir) / "pl_export"
    out = run(["gog", "drive", "download", file_id, "--out", str(dest)])
    for line in out.splitlines():
        if line.startswith("path\t"):
            return Path(line.split("\t", 1)[1].strip())
    raise RuntimeError(f"could not determine downloaded path from: {out[:200]}")


def a1_col(idx):
    """0-based column index -> A1 letters (0 -> A, 26 -> AA)."""
    col, letters = idx + 1, ""
    while col:
        col, rem = divmod(col - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def read_rows(path):
    path = Path(path)
    if path.suffix.lower() == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            return list(csv.reader(f))
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl needed for .xlsx; export as CSV or pip install openpyxl")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return [["" if c.value is None else c.value for c in row] for row in ws.iter_rows()]


# ------------------------------------------------------------------------- main

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", help="local P&L export (.csv or .xlsx)")
    ap.add_argument("--drive-id", help="Drive file id of the P&L")
    ap.add_argument("--year", type=int)
    ap.add_argument("--month", type=int)
    ap.add_argument("--write", action="store_true", help="apply corrections")
    ap.add_argument("--json", help="write the full report as JSON")
    args = ap.parse_args()

    tmp = tempfile.mkdtemp(prefix="storepl-")
    if args.file:
        path, source = Path(args.file), str(args.file)
    else:
        fid = args.drive_id
        source = f"Drive {fid}"
        if not fid:
            found = find_drive_pl(args.year, args.month)
            if not found:
                print("No monthly P&L found in Drive (looked for 'P&L <Month>_<Year>').")
                return 1
            fid, name, _ = found
            source = f"Drive: {name}"
            print(f"Using {source}")
        path = download_pl(fid, tmp)

    rows = read_rows(path)
    try:
        pl = sp.extract(rows)
    except ValueError as e:
        print(f"REFUSED: {e}")
        return 2

    year, month = pl["month"]
    start, end = pl["period"]
    if year < MIN_YEAR:
        print(f"REFUSED: P&L is {year}; automation floor is {MIN_YEAR} "
              "(earlier years are reconciled history).")
        return 2

    row_label = f"{month}.{year}"
    print(f"P&L window {start} .. {end}  ->  store-tab row '{row_label}'")
    print(f"Mode: {'WRITE' if args.write else 'DRY RUN'}\n")

    sheets_io.load_env()
    svc = sheets_io.sheets_service()
    sid = sheets_io.spreadsheet_id()
    report = {"source": source, "period": [str(start), str(end)],
              "row": row_label, "write": args.write, "stores": {}}
    updates = []
    totals = {}

    for tab, pl_store in sorted(pl["stores"].items()):
        vals = svc.spreadsheets().values().get(
            spreadsheetId=sid, range=f"'{tab}'!A1:BZ200").execute().get("values", [])
        target_idx, sheet_row = None, None
        for i, r in enumerate(vals):
            if r and str(r[0]).strip() == row_label:
                target_idx, sheet_row = i + 1, r
                break
        if sheet_row is None:
            print(f"{tab}: no row '{row_label}' -- skipped")
            report["stores"][tab] = {"error": f"no row {row_label}"}
            continue

        findings = sp.compare_row(pl_store, sheet_row, allow_overwrite=args.write)
        report["stores"][tab] = {"row_number": target_idx, "findings": findings}
        for f in findings:
            totals[f["action"]] = totals.get(f["action"], 0) + 1

        print(f"--- {tab.strip()} (row {target_idx}) ---")
        for f in findings:
            cur = "blank" if f["sheet"] is None else f"{f['sheet']:,.0f}"
            delta = "" if f["sheet"] in (None, 0) else f"  ({(f['pl']/f['sheet']-1)*100:+.1f}%)"
            flag = {"agree": "ok  ", "fill": "FILL", "update": "SET ", "report": "DIFF"}[f["action"]]
            print(f"  {flag} {f['name']:<18} P&L {f['pl']:>10,.0f}   sheet {cur:>10}{delta}")
            if f["action"] in ("fill", "update") and f["col"] is not None:
                updates.append({"range": f"'{tab}'!{a1_col(f['col'])}{target_idx}",
                                "values": [[f["pl"]]]})
        print()

    print(f"Summary: " + ", ".join(f"{k}={v}" for k, v in sorted(totals.items())))
    print(f"Cells to change: {len(updates)}")

    if args.json:
        Path(args.json).write_text(json.dumps(report, indent=2, default=str))
        print(f"Report -> {args.json}")

    if updates and args.write:
        svc.spreadsheets().values().batchUpdate(
            spreadsheetId=sid,
            body={"valueInputOption": "USER_ENTERED", "data": updates}).execute()
        print(f"WROTE {len(updates)} cells.")
    elif updates:
        print("Dry run -- nothing written. Re-run with --write to apply.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
